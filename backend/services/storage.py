"""
Simple file-based storage — saves ROI records to a local JSON file.
No database needed. File lives at backend/data/roi_records.json.

Each record carries a stable `record_id` used as the join key across the
export sheets (All_ROI_Data, SME_Audit_Log, Field_Provenance) and the
append-only audit log (see services/audit.py).
"""
import csv
import io
import json
import os
import uuid
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from models import ROIRecord
from services import audit

DATA_FILE = os.path.join(os.path.dirname(__file__), "..", "data", "roi_records.json")

# The 15 clean Domo columns (no provenance noise — that lives on its own sheet).
DOMO_COLUMNS = [
    "year", "client", "publisher", "date_delivered", "currency",
    "identified_risk", "id_cost_avoidance", "acc_cost_avoidance",
    "id_cost_optimization", "acc_cost_optimization", "realized_savings",
    "contract_spend", "pricing_available", "notes", "elevate_deliverable",
]

# Metric keys that can carry per-field provenance, with client-facing labels.
PROVENANCE_METRICS = [
    ("identified_risk",       "Identified Risk"),
    ("id_cost_avoidance",     "Identified Cost Avoidance"),
    ("acc_cost_avoidance",    "Accomplished Cost Avoidance"),
    ("id_cost_optimization",  "Identified Cost Optimization"),
    ("acc_cost_optimization", "Accomplished Cost Optimization"),
    ("realized_savings",      "Realized Savings"),
    ("contract_spend",        "Contract Spend"),
]


def _new_record_id() -> str:
    return f"r_{uuid.uuid4().hex[:12]}"


def _ensure_ids(records: list[dict]) -> bool:
    """Backfill a record_id onto any record missing one. Returns True if any
    record was changed (so the caller can persist the migration)."""
    changed = False
    for r in records:
        if not r.get("record_id"):
            r["record_id"] = _new_record_id()
            changed = True
    return changed


def _load() -> list[dict]:
    os.makedirs(os.path.dirname(DATA_FILE), exist_ok=True)
    if not os.path.exists(DATA_FILE):
        return []
    with open(DATA_FILE, "r") as f:
        records = json.load(f)
    # Migrate legacy records (saved before record_id existed) in place.
    if _ensure_ids(records):
        _save(records)
    return records


def _save(records: list[dict]):
    os.makedirs(os.path.dirname(DATA_FILE), exist_ok=True)
    with open(DATA_FILE, "w") as f:
        json.dump(records, f, indent=2, default=str)


def save_record(record: ROIRecord) -> dict:
    """Upsert an ROI record by source_file — prevents duplicates. Assigns a
    stable record_id (preserved across re-saves) and logs the store event."""
    records = _load()
    entry = record.model_dump()
    entry["saved_at"] = datetime.utcnow().isoformat()

    if entry.get("source_file"):
        for i, r in enumerate(records):
            if r.get("source_file") == entry["source_file"]:
                # Re-store of an existing file: keep its id, record an update event.
                entry["record_id"] = r.get("record_id") or _new_record_id()
                records[i] = entry
                _save(records)
                audit.append_event(
                    entry["record_id"], "update", user=entry.get("sme"),
                    note="Re-stored from extraction",
                )
                return entry

    # Brand-new record.
    entry["record_id"] = entry.get("record_id") or _new_record_id()
    records.append(entry)
    _save(records)
    audit.append_event(
        entry["record_id"], "create", user=entry.get("sme"),
        note="Stored from extraction",
    )
    return entry


def update_record(record_id: str, changes: dict, user: str | None = None,
                  note: str | None = None) -> dict:
    """Apply a partial edit to a stored record, persist it, and append one
    immutable audit event per changed field. Raises KeyError if not found."""
    records = _load()
    for r in records:
        if r.get("record_id") == record_id:
            for field, new_value in changes.items():
                if field in ("record_id", "saved_at"):
                    continue  # never editable
                old_value = r.get(field)
                if old_value == new_value:
                    continue  # no-op, don't log
                r[field] = new_value
                audit.append_event(
                    record_id, "edit", user=user, field=field,
                    old_value=old_value, new_value=new_value, note=note,
                )
            r["updated_at"] = datetime.utcnow().isoformat()
            _save(records)
            return r
    raise KeyError(record_id)


def get_all_records() -> list[dict]:
    return _load()


def export_csv() -> str:
    """Return all records as a CSV string with the 15 Domo columns."""
    records = _load()
    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=["record_id"] + DOMO_COLUMNS,
        extrasaction="ignore",
        lineterminator="\n",
    )
    writer.writeheader()
    for r in records:
        writer.writerow(r)
    return output.getvalue()


def _style_header(ws, columns, header_fill, header_font):
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autowidth(ws, columns):
    for col_idx, col_name in enumerate(columns, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(
            len(str(col_name)) + 2, 12
        )


def export_xlsx() -> bytes:
    """Return all records as an XLSX workbook with three sheets:
      1. All_ROI_Data    — clean ROI values + record_id (Domo-ready)
      2. SME_Audit_Log   — the append-only event history (creates + edits)
      3. Field_Provenance — per-field source slide + confidence (long format)
    All three join on record_id.
    """
    records = _load()
    by_id = {r.get("record_id"): r for r in records}
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    wb = Workbook()

    # ─── Sheet 1: All_ROI_Data (clean) ───────────────────────────────────────
    ws_data = wb.active
    ws_data.title = "All_ROI_Data"
    data_columns = (
        ["record_id"] + DOMO_COLUMNS
        + ["confidence", "source_file", "sme", "stored_name", "saved_at"]
    )
    _style_header(ws_data, data_columns, header_fill, header_font)
    for row_idx, record in enumerate(records, start=2):
        for col_idx, col_name in enumerate(data_columns, start=1):
            ws_data.cell(row=row_idx, column=col_idx, value=record.get(col_name))
    _autowidth(ws_data, data_columns)

    # ─── Sheet 2: SME_Audit_Log (event history) ──────────────────────────────
    ws_audit = wb.create_sheet("SME_Audit_Log")
    audit_columns = [
        "timestamp", "record_id", "client", "publisher", "year",
        "user", "action", "field", "old_value", "new_value", "note",
    ]
    _style_header(ws_audit, audit_columns, header_fill, header_font)

    events = audit.get_events()
    # Legacy fallback: records that predate the event log get a synthetic
    # "create" row so the sheet still reflects them.
    logged_ids = {e.get("record_id") for e in events}
    synthetic = [
        {
            "timestamp": r.get("saved_at"), "record_id": r.get("record_id"),
            "user": r.get("sme"), "action": "create", "field": None,
            "old_value": None, "new_value": None, "note": r.get("notes"),
        }
        for r in records if r.get("record_id") not in logged_ids
    ]
    rows = sorted(events + synthetic, key=lambda e: e.get("timestamp") or "")
    for row_idx, e in enumerate(rows, start=2):
        rec = by_id.get(e.get("record_id"), {})
        ws_audit.cell(row=row_idx, column=1, value=e.get("timestamp"))
        ws_audit.cell(row=row_idx, column=2, value=e.get("record_id"))
        ws_audit.cell(row=row_idx, column=3, value=rec.get("client"))
        ws_audit.cell(row=row_idx, column=4, value=rec.get("publisher"))
        ws_audit.cell(row=row_idx, column=5, value=rec.get("year"))
        ws_audit.cell(row=row_idx, column=6, value=e.get("user"))
        ws_audit.cell(row=row_idx, column=7, value=e.get("action"))
        ws_audit.cell(row=row_idx, column=8, value=e.get("field"))
        ws_audit.cell(row=row_idx, column=9, value=e.get("old_value"))
        ws_audit.cell(row=row_idx, column=10, value=e.get("new_value"))
        ws_audit.cell(row=row_idx, column=11, value=e.get("note"))
    _autowidth(ws_audit, audit_columns)

    # ─── Sheet 3: Field_Provenance (long format) ──────────────────────────────
    ws_prov = wb.create_sheet("Field_Provenance")
    prov_columns = [
        "record_id", "client", "publisher", "year",
        "metric", "value", "source_slide", "confidence", "alternates",
    ]
    _style_header(ws_prov, prov_columns, header_fill, header_font)
    prov_row = 2
    for record in records:
        fmeta = record.get("field_meta") or {}
        for key, label in PROVENANCE_METRICS:
            meta = fmeta.get(key)
            # Emit a row when there's a value or any provenance for this metric.
            if record.get(key) is None and not meta:
                continue
            meta = meta or {}
            alts = meta.get("alternates") or []
            alts_str = "; ".join(
                f"{a.get('value')} ({a.get('confidence')}%)" for a in alts
            ) if alts else None
            ws_prov.cell(row=prov_row, column=1, value=record.get("record_id"))
            ws_prov.cell(row=prov_row, column=2, value=record.get("client"))
            ws_prov.cell(row=prov_row, column=3, value=record.get("publisher"))
            ws_prov.cell(row=prov_row, column=4, value=record.get("year"))
            ws_prov.cell(row=prov_row, column=5, value=label)
            ws_prov.cell(row=prov_row, column=6, value=record.get(key))
            ws_prov.cell(row=prov_row, column=7, value=meta.get("source_slide"))
            ws_prov.cell(row=prov_row, column=8, value=meta.get("confidence"))
            ws_prov.cell(row=prov_row, column=9, value=alts_str)
            prov_row += 1
    _autowidth(ws_prov, prov_columns)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
